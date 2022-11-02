import jieba
import re
import math

import snownlp
from openpyxl import load_workbook
'''
功能：根据输入的字符串数组来进行计算情感值
使用方法：
在def calculation(txt)函数中 输入要分析的字符串数组
功能：会返回计算的整个数组的情感值

在def singleSentiment(text) 输入要分析的字符串数组
功能：会返回一个每个句子的情感子的数组

'''
negative_file = '否定词.txt'
adv_file = '程度副词.xlsx'
emotional_file = '情感词汇本体.xlsx'
test_file = '测试文章.txt'
emotion={
    '[尴尬]': -1 ,
    '[捂脸]':  -1,
    '[视频卫星]': 0 ,
    '[给心心]':  1,
    '[睡]':  0,
    '[奶瓶]': 0 ,
    '[打脸]': -1 ,
    '[丘比特]': 1 ,
    '[互粉]':  1,
    '[吃瓜群众]': 0 ,
    '[怪我咯]': 1 ,
    '[大笑]':  1,
    '[保卫萝卜_问号]': 0 ,
    '[击掌]':  1,
    '[灵魂出窍]': 0 ,
    '[发]':  1,
    '[困]':  -1,
    '[辣眼睛]':  -1,
    '[害羞]':  0,
    '[摊手]':  -1,
    '[酷]':  1,
    '[爱心]': 1 ,
    '[汤姆]':  0,
    '[酸]':  0,
    '[咖啡]':  0,
    '[保卫萝卜_笔芯]': 0 ,
    '[便便]':  -1,
    '[顶]':  1,
    '[嘿哈]':  1,
    '[挤眼]':  0,
    '[新蝙蝠侠]':  0,
    '[色]':  1,
    '[吐]':  -1,
    '[我想静静]': 0 ,
    '[烟花]':  1,
    '[强壮]':  1,
    '[哭泣]':  -1,
    '[怒]':  -1,
    '[许愿虎]': 0 ,
    '[亲亲]':  1,
    '[吃瓜]':  0,
    '[嗅嗅]':  0,
    '[笑哭]':  0,
    '[放假]':  0,
    '[不失礼貌的微笑]':  1,
    '[来古-注意]':  0,
    '[嘻嘻]':  1,
    '[12周年]':  1,
    '[勾引]':  1,
    '[原神_哼]': 0 ,
    '[馋嘴]':  1,
    '[酷拽]':  1,
    '[衰]':  -1,
    '[喜极而泣]': 1 ,
    '[吐彩虹]':  1,
    '[暴发虎]':  0,
    '[胜利]':  1,
    '[脱单doge]':  0,
    '[原神_喝茶]':  0,
    '[滑稽]':  1,
    '[2022]':  0,
    '[惊喜]':  1,
    '[V5]':  0,
    '[去污粉]': 0 ,
    '[裂开]':  -1,
    '[弱]':  -1,
    '[歪嘴]':  -1,
    '[坎公骑冠剑_吃鸡]': 0 ,
    '[坎公骑冠剑_无语]':  0,
    '[送花花]':  1,
    '[思考]':  0,
    '[心碎]':  -1,
    '[飞吻]':  1,
    '[嘘声]':  -1,
    '[鄙视]':  -1,
    '[鲜花]':  1,
    '[吐血]':  -1,
    '[红脸]':  0,
    '[拳头]':  -1,
    '[锦鲤]':  1,
    '[点赞]':  1,
    '[老鼠]':  0,
    '[擦汗]':  -1,
    '[酸了]':  0,
    '[右哼哼]':  -1,
    '[脸红]':  1,
    '[右边]':  0,
    '[冷]':  0,
    '[左哼哼]': -1 ,
    '[憨笑]':  1,
    '[呆无辜]':  0,
    '[给力]':  1,
    '[奸笑]':  -1,
    '[熊吉]':  0,
    '[福到了]': 1 ,
    '[哈哈]':  1,
    '[赢牛奶]': 0 ,
    '[红包]':  0,
    '[抓狂]':  -1,
    '[坎公骑冠剑_钻石]': 0 ,
    '[偷笑]':  1,
    '[保卫萝卜_白眼]': -1 ,
    '[尬笑]':  0,
    '[OK]':  1,
    '[藏狐]':  0,
    '[保卫萝卜_哭哭]':  0,
    '[报税]':  0,
    '[感冒]':  0,
    '[彩虹屁]': 0 ,
    '[小鼓掌]':  1,
    '[傲娇]':  1,
    '[翻白眼]': -1 ,
    '[虎年]':  0,
    '[灵机一动]': 1 ,
    '[来古-沉思]': 0 ,
    '[打call]':  1,
    '[保佑]':  0,
    '[奋斗]':  1,
    '[可爱]':  1,
    '[泼水节]':  0,
    '[红灯笼]':  0,
    '[胡瓜]':  0,
    '[捂眼]':  0,
    '[骷髅]':  -1,
    '[强]':  1,
    '[敲打]':  -1,
    '[哈欠]':  -1,
    '[小雪人]':  0,
    '[悠闲]':  1,
    '[菜刀]':  0,
    '[鸡腿]':  0,
    '[摸头]':  0,
    '[拜拜]':  0,
    '[猪头]':  0,
    '[邓布利多]': 0 ,
    '[坏笑]':  -1,
    '[太开心]': 1 ,
    '[水稻]':  0,
    '[可怜]':  -1,
    '[玫瑰]':  1,
    '[做鬼脸]':  0,
    '[牛年]':  0,
    '[交税]':  0,
    '[抱拳]':  0,
    '[闭嘴]':  -1,
    '[碰拳]':  1,
    '[悲伤]':  -1,
    '[666]':  1,
    '[呆]':  0,
    '[派对]': 0 ,
    '[羞羞]':  0,
    '[保卫萝卜_哇]': 0 ,
    '[挖鼻]':  -1,
    '[流泪]':  0,
    '[泪奔]':  -1,
    '[机智]':  1,
    '[惊讶]':  0,
    '[钱]':  0,
    '[握手]':  1,
    '[雪花]':  0,
    '[紫薇别走]': 0 ,
    '[来古-呆滞]':  0,
    '[原神_欸嘿]':  0,
    '[墨镜]':  1,
    '[哦呼]':  1,
    '[无语]':  -1,
    '[格林德沃]':  0,
    '[笑]':  1,
    '[汗]':  0,
    '[汤圆]':  0,
    '[苦涩]':  -1,
    '[耶]':  1,
    '[单身奖杯]': 0 ,
    '[开学季]':  0,
    '[黑洞]':  0,
    '[虎爪比心]': 0 ,
    '[撇嘴]':  -1,
    '[得意]':  1,
    '[怒骂]':  -1,
    '[皱眉]':  -1,
    '[哼]':  -1,
    '[支持]': 1 ,
    '[笑而不语]': 0 ,
    '[绝望的凝视]': 0 ,
    '[疑问]':  0,
    '[流汗]':  -1,
    '[黑脸]':  -1,
    '[抱抱]':  1,
    '[舔屏]':  1,
    '[来看我]': 0 ,
    '[恐惧]':  -1,
    '[狗子]':  0,
    '[萌虎贴贴]': 0 ,
    '[听歌]':  0,
    '[再见]':  0,
    '[干杯]':  1,
    '[惊恐]':  -1,
    '[求饶]':  -1,
    '[举手]':  0,
    '[绿马]':  0,
    '[吓]':  -1,
    '[星星眼]':  1,
    '[阴险]':  -1,
    '[跪了]':  0,
    '[加油]':  1,
    '[春游家族]': 0 ,
    '[呲牙]':  1,
    '[18禁]':  -1,
    '[嗑瓜子]':  0,
    '[快哭了]':  -1,
    '[来古-震撼]':  0,
    '[月亮]':  0,
    '[妙啊]':  1,
    '[原神_生气]': -1 ,
    '[愉快]':  1,
    '[杰瑞]':  0,
    '[如花]':  0,
    '[喜欢]':  1,
    '[谜语人]':  0,
    '[太阳]':  0,
    '[感谢]':  1,
    '[调皮]':  1,
    '[嘟嘟]':  1,
    '[疑惑]':  0,
    '[晕]':  0,
    '[泣不成声]':  -1,
    '[抠鼻]':  0,
    '[航天员]':  0,
    '[看]':  0,
    '[awsl]':  0,
    '[绿帽子]':  0,
    '[比心]':  1,
    '[嫌弃]':  -1,
    '[咒骂]':  -1,
    '[小丑]':  0,
    '[委屈]':  -1,
    '[礼物]':  0,
    '[口罩]':  0,
    '[黑线]':  -1,
    '[左上]':  0,
    '[加好友]': 0 ,
    '[弹幕破百亿]':  0,
    '[单身狗]':  0,
    '[并不简单]': 0 ,
    '[豹富]':  0,
    '[二哈]':  0,
    '[疼]':  -1,
    '[笑cry]': 1 ,
    '[斜眼]':  1,
    '[doge]':  1,
    '[吃惊]':  0,
    '[微笑]':  1,
    '[哇]':  1,
    '[发怒]': -1 ,
    '[啤酒]':  0,
    '[生气]':  -1,
    '[揣手]':  0,
    '[爱你]':  1,
    '[白眼]':  -1,
    '[凋谢]':  0,
    '[污]':  0,
    '[傻眼]': -1 ,
    '[大哭]':  -1,
    '[生病]':  -1,
    '[来古-疑问]': 0 ,
    '[喵喵]':  0,
    '[囧]':  0,
    '[拥抱]': 1 ,
    '[泪]':  -1,
    '[送心]':  1,
    '[傲慢]':  -1,
    '[响指]':  0,
    '[石化]':  0,
    '[原神_哇]':0  ,
    '[赞]':  1,
    '[炸弹]': -1 ,
    '[西瓜]':  0,
    '[吐舌]':  1,
    '[福气虎]':0  ,
    '[失望]':  -1,
    '[震惊]':  0,
    '[鼓掌]':  1,
    '[允悲]':  -1,
    '[嘘]':  0,
    '[原神_嗯]': 0 ,
    '[憧憬]':  1,
    '[费解]':  -1,
    '[发呆]':  0,
    '[抱一抱]': 1 ,
    '[难过]':  -1,
    '[不看]':  -1,
    '[中国赞]':0
}
alpha = 0.75   # 不确定的情感值极性前后判断因素的比例，默认为0.5
dict_of_emtion={}
dict_of_adv={}
list_of_negative=[]
'''
情感计算流程

1.将语句分句

2.将句子分词

3.判断词语 格式[[[强度，极性]]，否定词，副词]
【添加：若句子中存在表情，正向表情则为[[[9,1]],1,1]]，负向表情则为[[[9,-1]],1,1]】
若词语只有一个情感极值
该词语情感值为 强度*极性*否定词*副词
【强度 0～9，极性 0 中性。1褒义。2贬义｛最后会处理成-1｝。3褒贬不一，否定词 1 无。-1有。】
若一个情感词有两种相同的极性且都是3，那么该情感值的极性由该情感词前0-4个的情感值极性*0.75(可以修改前后比重), 
后0-4个情感值极性*0.25的和共同确定，计算只包括0,1,-1。最后结果根据离-1，0,1这三个数的绝对距离确定，
若一个情感词前后极性不同，那么该情感值的极性同上。

4.最后的文本情感值为：
语句情感值相加/语句情感绝对值相加

eg：
[[[9,1]],1,1]
[[[7,-1]],1,1]
calculate：
(9-7)/(9+7)=0.125

'''
def analysisEmo(text):
    ''' 判断是否含有表情，返回1则为正向，-1为负向，0则无法判断，需要根据句子情感值判断 '''
    sentence_emo=[]
    n = re.findall(r"\[(.+?)\]", str(text))
    if len(n):
        for k in n:
            if '['+k+']' in emotion:
                sentence_emo.append(emotion['['+k+']'])
    finnal_emo=0
    for i in sentence_emo:
        finnal_emo+=i
    if finnal_emo>0:
        return 1
    if finnal_emo==0:
        return 0
    if finnal_emo<0:
        return -1
def anaysisPolarity(word, dict_of_emtion):
    str1 = dict_of_emtion[word][0][0]  # 第一个强度
    plo1 = dict_of_emtion[word][0][1]  # 第一个极性
    if len(dict_of_emtion[word]) > 1:   # 若有两个极性
        str2 = dict_of_emtion[word][1][0]
        plo2 = dict_of_emtion[word][1][1]
        if plo1 == plo2 and plo1 in [-1, 0, 1]:  # 判断依据1
            return [[str1, plo1]]            #  返回第一个强度，极性
        elif (plo1 == plo2 and plo1 == 3) or (plo1 != plo2):  # 判断依据2,3
            return [[str1, plo1], [str2, plo2]]         # 两个都返回
    else:
        return [[str1, plo1]]

'''传入参数：文段，情感字典，副词字典，否定词列表'''

def analysisWords(words, dict_of_emtion, dict_of_adv, list_of_negative, par_W):
    for word in words:
        if word in dict_of_emtion.keys():  # 如果这个词在情感词中，则进行分析
            w3 = 1  # 默认没有否定词
            w4 = 1  # 默认副词为没有，也就是弱，为1
            w1w2 = anaysisPolarity(word, dict_of_emtion)  # 判断极性
            for num in range(1, words.index(word)):
                index = words.index(word) - num
                index_w = words[index]  # 当前下标表示的词语
                if index_w == '，':
                    break
                else:
                    if index_w in list_of_negative:  # 如果在否定词列表中
                        w3 *= -1  # 找到了否定词，置为-1
                    if index_w in dict_of_adv.keys():
                        w4 = dict_of_adv[index_w]  # 副词
            try:
                par_W.append([w1w2, w3, w4])
            except Exception as e:
                print("错误:", e)

# PA,PE, 乐
# PD,PH,PG,PB,PK,NA 好
# NA 怒
# NB,NJ,NH,PF 哀
# NI,NC,NG 惧
# NE,ND,NN,NK,NL 恶
# PC 惊

def distanceOfNum(result, x1, x2):
    num1 = math.fabs(result - x1)
    num2 = math.fabs(result - x2)
    if num1 > num2:
        return x2
    else:
        return x1
def readWord():
    '''读取情感词汇到字典'''
    wb = load_workbook(emotional_file)
    ws = wb[wb.sheetnames[0]]  # 读取第一个sheet
    for i in range(2, ws.max_row):
        word = ws['A' + str(i)].value#读取词
        strength = ws['F' + str(i)].value  # 一个强度
        polarity = ws['G' + str(i)].value  # 一个极性
        if polarity == 2:#2为否定词
            polarity = -1
        assist = ws['H' + str(i)].value  # 辅助情感分类
        if word not in dict_of_emtion.keys():
            dict_of_emtion[word] = list([[strength, polarity]])
        else:
            dict_of_emtion[word].append([strength, polarity])  # 添加二义性的感情词
        if assist != None:
            str2 = ws['I' + str(i)].value #  获得含有辅助情感词的强度和极性
            pola2 = ws['J' + str(i)].value
            if pola2 == 2:
                pola2 = -1
            dict_of_emtion[word].append([str2, pola2])
def readadv():
    '''读取程度副词副字典'''
    wb = load_workbook(adv_file)
    ws = wb[wb.sheetnames[0]]
    for i in range(2, ws.max_row):
        dict_of_adv[ws['A' + str(i)].value] = ws['B' + str(i)].value #提取程度副词
def readNegative():
    '''读取否定词列表'''
    with open(negative_file, "r", encoding='utf-8') as f:
        temps = f.readlines()
    for temp in temps:
        list_of_negative.append(temp.replace("\n", "")) #提取否定词
#para_w= [[强度,极性]、是否否定词、副词]
def calSim(txt):
    para_W = []  # 词语的极性，副词和否定词
    for info in txt:
        # article = re.findall(r'[^。！？\s]+', info)  # 一句一句分析 返回string中所有与pattern匹配的全部字符串,返回形式为数组。
        article = info # 一句一句分析 返回string中所有与pattern匹配的全部字符串,返回形式为数组。
        if len(article) != 0:
            bol=analysisEmo(article)
            if bol == 1:
                para_W.append([[[9, 1]], 1, 1])
            elif bol == -1:
                para_W.append([[[9, -1]], 1, 1])
            elif bol == 0:
                words = jieba.lcut(article)  # 一句一句分词
                analysisWords(words, dict_of_emtion, dict_of_adv, list_of_negative, para_W)
    new_para_W = {}
    index = 1  # 情感参数的数量
    for x in para_W:  # 将只有一个情感极值的列表合并
        if len(x[0]) == 1:
            w = x[0][0][0] * x[0][0][1]
            for numx in x[1:]:
                w *= numx
            new_para_W[index] = w #得出只有一个情感极值的词语的最终情感值
        else:
            new_para_W[index] = x
        index += 1
    for i in range(1, len(new_para_W) + 1):
        if type(new_para_W[i]) == list:  # 如果该情感值是未计算的
            temp_result = 0
            k = i-1 if i-1 != 0 else i
            index = 1
            while index <= 4:  #  计算0-4个的值
                if type(new_para_W[k]) != list:
                    temp_result += new_para_W[k] * alpha   # 当前值乘以alaph
                else:
                    temp_result += new_para_W[k][0][0][1] * alpha  #如果没有，则默认为第一个极性
                k -= 1
                index += 1
                if k <= 0:
                    break
            if i+1 < len(new_para_W):
                k = i + 1
            else:
                k=i  #  计算后四个的值
            index = 1
            while index <= 4:
                if type(new_para_W[k]) != list and new_para_W[k] != 3:  # 只考虑后面
                    temp_result += new_para_W[k] * (1-alpha)
                k += 1
                index += 1
                if k > len(new_para_W):  # 如果超出了最长长度，后面则不考虑计算
                    break
            w2 = distanceOfNum(temp_result, new_para_W[i][0][0][1], new_para_W[i][0][1][1])  # 求出w2
            dict_of_str_plo = {}  # 存放极值---强度的字典
            str1 = new_para_W[i][0][0][0]
            if new_para_W[i][0][0][1] == 3:
                plo1 = w2
            else:
                plo1 = new_para_W[i][0][0][1]  # 将褒贬不一的置为求出的局部感情极值
            str2 = new_para_W[i][0][1][0]
            if new_para_W[i][0][1][1] == 3 :
                plo2 = w2
            else:
                plo2=new_para_W[i][0][1][1]
            dict_of_str_plo[plo2] = str2
            dict_of_str_plo[plo1] = str1
            if w2 == 0:
                new_para_W[i] = 0
            else:
                try:
                    new_para_W[i] = dict_of_str_plo[w2] * new_para_W[i][1] * new_para_W[i][2]
                except Exception as e:
                    print("错误：", e)
    molecular = 0   # 分子
    denominator = 0  # 分母
    for value in new_para_W.values():
        molecular += value
        denominator += math.fabs(value)
    if denominator == 0:
        return 0.0
    else:
        return (molecular/denominator)
def singleSentiment(text):
    Ssenti = []
    for w in text:
        sent=(snownlp.SnowNLP(w).sentiments-0.5)*2
        Ssenti.append(sent);
        print(w+"：的情感值为："+str(sent))
    return Ssenti
def calculation(txt):
    readWord()
    readadv()
    readNegative()
    totalSenti=calSim(txt)
    return totalSenti
